# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'G:\设备数据分析\城西闪传_修复版_更新后.xlsx', read_only=True)
output = []

output.append('=== Sheets ===')
for sheet_name in wb.sheetnames:
    output.append(f'  {sheet_name}')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    output.append(f'\n=== {sheet_name} ===')
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
    output.append(f'Columns: {len(headers)}')
    for i, h in enumerate(headers):
        output.append(f'  [{i}] {h}')
    
    row_count = ws.max_row - 1
    output.append(f'Data rows: {row_count}')
    
    output.append('Sample data (rows 4-6):')
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=6, values_only=True)):
        output.append(f'  Row {i+4}: {list(row)}')
wb.close()

with open(r'G:\设备数据分析\xlsx_info.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))
print('Done - wrote to xlsx_info.txt')
