import os
import math
import logging
import openpyxl
from utils.coord_transform import cgcs2000_to_wgs84, cgcs2000_to_gcj02
import config

logger = logging.getLogger(__name__)

# REST API fallback
REST_URL = os.environ.get('SUPABASE_URL', '')
REST_KEY = os.environ.get('SUPABASE_KEY', '')
REST_HEADERS = {'apikey': REST_KEY, 'Authorization': f'Bearer {REST_KEY}'}


def _rest_get(table, query=''):
    import requests
    url = f'{REST_URL}/rest/v1/{table}'
    if query:
        url += f'?{query}'
    try:
        r = requests.get(url, headers=REST_HEADERS, timeout=30)
        r.raise_for_status()
        return r.json()
    except requests.exceptions.SSLError:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        r = requests.get(url, headers=REST_HEADERS, timeout=30, verify=False)
        r.raise_for_status()
        return r.json()


def _rest_upsert(table, data):
    """Insert or update a single record via Supabase REST API"""
    import requests
    url = f'{REST_URL}/rest/v1/{table}'
    headers = {**REST_HEADERS, 'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates'}
    try:
        r = requests.post(url, json=data, headers=headers, timeout=30)
        r.raise_for_status()
        return True
    except Exception as e:
        logger.warning(f"REST upsert to {table} failed: {e}")
        return False


def _rest_upsert_batch(table, records):
    """Insert or update multiple records via Supabase REST API"""
    import requests
    if not records:
        return 0
    url = f'{REST_URL}/rest/v1/{table}'
    headers = {**REST_HEADERS, 'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates'}
    try:
        r = requests.post(url, json=records, headers=headers, timeout=60)
        r.raise_for_status()
        return len(records)
    except Exception as e:
        logger.warning(f"REST batch upsert to {table} failed: {e}")
        return 0


def _get_conn():
    """获取数据库连接（PostgreSQL 或 SQLite），PostgreSQL 失败时降级到 SQLite"""
    if config.DB_TYPE == 'postgresql':
        try:
            import psycopg2
            import psycopg2.extras
            conn = psycopg2.connect(config.DATABASE_URL)
            conn.autocommit = False
            return conn
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(
                "PostgreSQL connection failed (%s), falling back to SQLite", e)
            # 降级时更新全局数据库类型，确保后续 SQL 查询使用正确的语法
            config.DB_TYPE = 'sqlite'
    import sqlite3
    conn = sqlite3.connect(config.DB_PATH, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.row_factory = sqlite3.Row
    return conn


def _fetch_all(cursor, conn):
    """获取所有结果（兼容 SQLite 和 PostgreSQL）"""
    if config.DB_TYPE == 'postgresql':
        cols = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        return [dict(zip(cols, row)) for row in rows]
    else:
        return [dict(row) for row in cursor.fetchall()]


def _sql(sqlite_sql, pg_sql=None):
    """根据数据库类型返回对应的 SQL"""
    if config.DB_TYPE == 'postgresql' and pg_sql:
        return pg_sql
    return sqlite_sql


def _hours_ago(hours):
    """返回 N 小时前的时间条件"""
    if config.DB_TYPE == 'postgresql':
        return f"NOW() - INTERVAL '{hours} hours'"
    else:
        return f"datetime('now', '-{hours} hours', 'localtime')"


def load_devices():
    from utils.coord_transform import wgs84_to_gcj02
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT device_id, name, area_name, device_type, manufacturers, address,
                   longitude, latitude, first_seen, last_seen
            FROM devices
        """)
        devices = []
        for row in _fetch_all(cursor, conn):
            lon = row['longitude']
            lat = row['latitude']
            if lon and lat:
                lon, lat = wgs84_to_gcj02(lon, lat)
            devices.append({
                'device_id': row['device_id'],
                'name': row['name'],
                'area_name': row['area_name'],
                'device_type': row['device_type'],
                'manufacturers': row['manufacturers'],
                'address': row['address'],
                'longitude': lon,
                'latitude': lat,
                'first_seen': row['first_seen'],
                'last_seen': row['last_seen']
            })
        conn.close()
        return devices
    except Exception as e:
        logger.warning(f"DB connection failed, using REST API: {e}")
        rows = _rest_get('devices', 'select=device_id,name,area_name,device_type,manufacturers,address,longitude,latitude,first_seen,last_seen')
        devices = []
        for row in rows:
            lon = row.get('longitude')
            lat = row.get('latitude')
            if lon and lat:
                lon, lat = wgs84_to_gcj02(lon, lat)
            devices.append({
                'device_id': row['device_id'],
                'name': row.get('name'),
                'area_name': row.get('area_name'),
                'device_type': row.get('device_type'),
                'manufacturers': row.get('manufacturers'),
                'address': row.get('address'),
                'longitude': lon,
                'latitude': lat,
                'first_seen': row.get('first_seen'),
                'last_seen': row.get('last_seen')
            })
        return devices


def load_pipe_nodes():
    wb = openpyxl.load_workbook(config.XLSX_PATH, read_only=True, data_only=True)
    ws = wb['管点']
    nodes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        point_id = row[0]
        x_val = row[16]
        y_val = row[17]
        pipe_type = row[9] if len(row) > 9 else None
        sub_type = row[10] if len(row) > 10 else None
        feature = row[15] if len(row) > 15 else None
        ground_elev = row[18] if len(row) > 18 else None
        well_bottom_elev = row[19] if len(row) > 19 else None
        depth = row[11] if len(row) > 11 else None

        if point_id and x_val and y_val:
            try:
                x = float(x_val)
                y = float(y_val)
                lon, lat = cgcs2000_to_gcj02(x, y)

                def safe_float(v):
                    if v is None:
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                nodes.append({
                    'point_id': str(point_id),
                    'pipe_type': str(pipe_type) if pipe_type else '',
                    'sub_type': str(sub_type) if sub_type else '',
                    'feature': str(feature) if feature else '',
                    'ground_elev': safe_float(ground_elev),
                    'well_bottom_elev': safe_float(well_bottom_elev),
                    'depth': safe_float(depth),
                    'x_cgcs': x,
                    'y_cgcs': y,
                    'lon': lon,
                    'lat': lat
                })
            except (ValueError, TypeError):
                continue
    wb.close()
    return nodes


def load_pipes():
    wb = openpyxl.load_workbook(config.XLSX_PATH, read_only=True, data_only=True)
    ws = wb['管线']
    pipes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        pipe_type = row[1] if len(row) > 1 else None
        sub_type = row[2] if len(row) > 2 else None
        start_id = row[9] if len(row) > 9 else None
        end_id = row[10] if len(row) > 10 else None
        diameter = row[20] if len(row) > 20 else None

        if start_id and end_id:
            pipes.append({
                'pipe_type': str(pipe_type) if pipe_type else '',
                'sub_type': str(sub_type) if sub_type else '',
                'start_id': str(start_id),
                'end_id': str(end_id),
                'diameter': str(diameter) if diameter else ''
            })
    wb.close()
    return pipes


def bind_devices_to_nodes(devices, nodes):
    if not nodes:
        return [{**d, 'bound_node': None, 'distance': None} for d in devices]

    bindings = []
    for dev in devices:
        if dev['longitude'] is None or dev['latitude'] is None:
            bindings.append({**dev, 'bound_node': None, 'distance': None})
            continue

        # 暴力搜索最近管点（纯Python，无numpy依赖）
        best_idx, best_dist = None, float('inf')
        for i, n in enumerate(nodes):
            d = haversine_distance(dev['longitude'], dev['latitude'], n['lon'], n['lat'])
            if d < best_dist:
                best_dist = d
                best_idx = i

        if best_dist > config.BIND_DISTANCE_MAX:
            bindings.append({**dev, 'bound_node': None, 'distance': round(best_dist, 2)})
        else:
            bindings.append({
                **dev,
                'bound_node': nodes[best_idx],
                'distance': round(best_dist, 2)
            })
    return bindings


def haversine_distance(lon1, lat1, lon2, lat2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def build_topology(nodes, pipes):
    node_map = {n['point_id']: n for n in nodes}
    adjacency = {}
    for pipe in pipes:
        s, e = pipe['start_id'], pipe['end_id']
        if s not in adjacency:
            adjacency[s] = []
        if e not in adjacency:
            adjacency[e] = []
        adjacency[s].append((e, pipe))
        adjacency[e].append((s, pipe))
    return node_map, adjacency


def get_device_readings(device_id, hours=168):
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        # 先获取连接（可能更新 DB_TYPE），再生成 SQL
        time_cond = _hours_ago(hours)
        sql = _sql(
            f"SELECT recorded_at, liquid_level, ammonia_n, cod, voltage, temperature, isonline FROM readings WHERE device_id = ? AND recorded_at >= {time_cond} ORDER BY recorded_at DESC",
            f"SELECT recorded_at, liquid_level, ammonia_n, cod, voltage, temperature, isonline FROM readings WHERE device_id = %s AND recorded_at >= {time_cond} ORDER BY recorded_at DESC"
        )
        if config.DB_TYPE == 'postgresql':
            cursor.execute(sql, (device_id,))
        else:
            cursor.execute(sql, (device_id,))
        rows = _fetch_all(cursor, conn)
        conn.close()
        return rows
    except Exception as e:
        logger.debug(f"DB failed for device_readings, using REST API: {e}")
        try:
            from datetime import datetime, timedelta
            cutoff = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
            return _rest_get('readings',
                f'select=recorded_at,liquid_level,ammonia_n,cod,voltage,temperature,isonline&device_id=eq.{device_id}&recorded_at=gte.{cutoff}&order=recorded_at.desc&limit=1000')
        except Exception:
            return []


def get_device_statistics(device_id, hours=48):
    """计算设备统计数据：均值、标准差、变化率、阈值状态"""
    readings = get_device_readings(device_id, hours=hours)
    if not readings:
        return None

    levels = [float(r['liquid_level']) for r in readings if r.get('liquid_level') is not None]
    if not levels:
        return None

    mean_level = sum(levels) / len(levels)
    std_level = (sum((x - mean_level) ** 2 for x in levels) / len(levels)) ** 0.5

    if len(levels) >= 2:
        change_rate = (levels[0] - levels[-1]) / len(levels)
    else:
        change_rate = 0

    if levels[0] > 3.0:
        status = "🔴 严重"
    elif levels[0] > 2.5:
        status = "🟠 高风险"
    elif levels[0] > 2.0:
        status = "🟡 偏高"
    else:
        status = "🟢 正常"

    if change_rate > 0.01:
        trend = "↑ 上升"
    elif change_rate < -0.01:
        trend = "↓ 下降"
    else:
        trend = "→ 稳定"

    return {
        'mean': round(mean_level, 3),
        'std': round(std_level, 3),
        'min': round(min(levels), 3),
        'max': round(max(levels), 3),
        'current': round(levels[0], 3),
        'change_rate': round(change_rate, 4),
        'status': status,
        'trend': trend,
        'count': len(levels),
    }


def get_all_recent_readings(hours=168):
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        time_cond = _hours_ago(hours)
        sql = _sql(
            f"SELECT device_id, recorded_at, liquid_level, ammonia_n, cod, voltage, temperature, isonline FROM readings WHERE recorded_at >= {time_cond} ORDER BY device_id, recorded_at DESC",
            f"SELECT device_id, recorded_at, liquid_level, ammonia_n, cod, voltage, temperature, isonline FROM readings WHERE recorded_at >= {time_cond} ORDER BY device_id, recorded_at DESC"
        )
        cursor.execute(sql)
        rows = _fetch_all(cursor, conn)
        conn.close()
        return rows
    except Exception as e:
        logger.warning(f"DB connection failed for readings, using REST API: {e}")
        from datetime import datetime, timedelta
        cutoff = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
        return _rest_get('readings',
            f'select=device_id,recorded_at,liquid_level,ammonia_n,cod,voltage,temperature,isonline&recorded_at=gte.{cutoff}&order=device_id,recorded_at.desc&limit=5000')


def get_device_online_stats():
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        sql = _sql("""
            SELECT r.device_id,
                   COUNT(*) as total_readings,
                   SUM(CASE WHEN r.isonline = 1 THEN 1 ELSE 0 END) as online_readings,
                   d.last_seen
            FROM readings r
            LEFT JOIN devices d ON r.device_id = d.device_id
            GROUP BY r.device_id
        """, """
            SELECT r.device_id,
                   COUNT(*) as total_readings,
                   SUM(CASE WHEN r.isonline = '1' THEN 1 ELSE 0 END) as online_readings,
                   d.last_seen
            FROM readings r
            LEFT JOIN devices d ON r.device_id = d.device_id
            GROUP BY r.device_id
        """)
        stats = {}
        for row in _fetch_all(cursor, conn):
            total = row['total_readings']
            online = row['online_readings'] or 0
            stats[row['device_id']] = {
                'total_readings': total,
                'online_readings': online,
                'online_rate': round(online / total * 100, 1) if total > 0 else 0,
                'last_seen': row['last_seen']
            }
        conn.close()
        return stats
    except Exception as e:
        logger.debug(f"DB failed for online_stats, using REST API: {e}")
        try:
            readings = _rest_get('readings', 'select=device_id,isonline,recorded_at&limit=10000')
            devices = _rest_get('devices', 'select=device_id,last_seen')
            dev_map = {d['device_id']: d.get('last_seen') for d in devices}
            stats = {}
            for r in readings:
                did = r.get('device_id')
                if did not in stats:
                    stats[did] = {'total_readings': 0, 'online_readings': 0}
                stats[did]['total_readings'] += 1
                if str(r.get('isonline')) == '1':
                    stats[did]['online_readings'] += 1
            for did, s in stats.items():
                t = s['total_readings']
                o = s['online_readings']
                s['online_rate'] = round(o / t * 100, 1) if t > 0 else 0
                s['last_seen'] = dev_map.get(did)
            return stats
        except Exception:
            return {}


def get_area_stats(bindings, readings, anomalies):
    area_stats = {}
    for b in bindings:
        area = b.get('area_name', '未知')
        if area not in area_stats:
            area_stats[area] = {
                'device_count': 0,
                'total_readings': 0,
                'anomaly_count': 0,
                'avg_liquid_level': 0,
                'liquid_levels': []
            }
        area_stats[area]['device_count'] += 1

    device_readings = {}
    for r in readings:
        did = r['device_id']
        if did not in device_readings:
            device_readings[did] = []
        device_readings[did].append(r)

    for b in bindings:
        area = b.get('area_name', '未知')
        did = b['device_id']
        if did in device_readings:
            area_stats[area]['total_readings'] += len(device_readings[did])
            for r in device_readings[did]:
                if r.get('liquid_level'):
                    area_stats[area]['liquid_levels'].append(r['liquid_level'])

    for a in anomalies.get('threshold', []):
        did = a['device_id']
        for b in bindings:
            if b['device_id'] == did:
                area = b.get('area_name', '未知')
                area_stats[area]['anomaly_count'] += 1
                break

    for area in area_stats:
        levels = area_stats[area]['liquid_levels']
        if levels:
            area_stats[area]['avg_liquid_level'] = round(sum(levels) / len(levels), 3)
        del area_stats[area]['liquid_levels']

    return area_stats


def get_manufacturer_stats(bindings, readings):
    mfr_stats = {}
    for b in bindings:
        mfr = b.get('manufacturers') or '未知'
        if mfr not in mfr_stats:
            mfr_stats[mfr] = {
                'device_count': 0,
                'devices': []
            }
        mfr_stats[mfr]['device_count'] += 1
        mfr_stats[mfr]['devices'].append(b['device_id'])

    return mfr_stats


def get_device_status(device_id, anomalies):
    status = {
        'overflow': 'normal',
        'fullness': 'normal',
        'overall': 'normal',
        'color': '#4caf50',
        'details': []
    }

    for a in anomalies.get('overflow', []):
        if a['device_id'] == device_id:
            if a['severity'] == 'critical':
                status['overflow'] = 'critical'
                status['color'] = '#d32f2f'
                status['details'].append(f"溢出风险: {a['overflow_risk']}m")
            elif a['severity'] == 'high' and status['overflow'] != 'critical':
                status['overflow'] = 'high'
                status['color'] = '#ff9800'
                status['details'].append(f"高溢出风险: {a['overflow_risk']}m")
            elif a['severity'] == 'medium' and status['overflow'] == 'normal':
                status['overflow'] = 'medium'
                status['color'] = '#ffc107'
                status['details'].append(f"中溢出风险: {a['overflow_risk']}m")

    for a in anomalies.get('pipe_fullness', []):
        if a['device_id'] == device_id:
            if a['severity'] == 'high':
                status['fullness'] = 'high'
                if status['color'] not in ['#d32f2f']:
                    status['color'] = '#d32f2f'
                status['details'].append(f"管道超载: {a['fullness']:.0%}")
            elif a['severity'] == 'medium' and status['fullness'] == 'normal':
                status['fullness'] = 'medium'
                if status['color'] not in ['#d32f2f', '#ff9800']:
                    status['color'] = '#ff9800'
                status['details'].append(f"高充满度: {a['fullness']:.0%}")

    if status['overflow'] == 'critical' or status['fullness'] == 'high':
        status['overall'] = 'critical'
    elif status['overflow'] in ['high', 'medium'] or status['fullness'] == 'medium':
        status['overall'] = 'warning'

    return status


def devices_to_geojson(bindings, anomalies=None, active_device_ids=None):
    features = []
    for b in bindings:
        if b.get('longitude') and b.get('latitude'):
            device_id = b['device_id']
            is_active = active_device_ids is None or device_id in active_device_ids
            status = get_device_status(device_id, anomalies) if anomalies else {
                'overall': 'normal',
                'color': '#4caf50',
                'details': []
            }

            if not is_active:
                status['overall'] = 'inactive'
                status['color'] = '#9e9e9e'
                status['details'] = ['离线/无数据']

            features.append({
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': [b['longitude'], b['latitude']]
                },
                'properties': {
                    'device_id': device_id,
                    'name': b.get('name', ''),
                    'device_type': b.get('device_type', ''),
                    'area_name': b.get('area_name', ''),
                    'bound_node_id': b['bound_node']['point_id'] if b.get('bound_node') else None,
                    'bind_distance': b.get('distance'),
                    'status': status['overall'],
                    'color': status['color'],
                    'details': status['details']
                }
            })
    return {'type': 'FeatureCollection', 'features': features}


def nodes_to_geojson(nodes):
    features = []
    for n in nodes:
        features.append({
            'type': 'Feature',
            'geometry': {
                'type': 'Point',
                'coordinates': [n['lon'], n['lat']]
            },
            'properties': {
                'point_id': n['point_id'],
                'pipe_type': n['pipe_type'],
                'sub_type': n['sub_type'],
                'feature': n['feature'],
                'ground_elev': n.get('ground_elev'),
                'well_bottom_elev': n.get('well_bottom_elev'),
                'depth': n.get('depth')
            }
        })
    return {'type': 'FeatureCollection', 'features': features}


def pipes_to_geojson(pipes, node_map):
    features = []
    for i, pipe in enumerate(pipes):
        start = node_map.get(pipe['start_id'])
        end = node_map.get(pipe['end_id'])
        if start and end:
            features.append({
                'type': 'Feature',
                'geometry': {
                    'type': 'LineString',
                    'coordinates': [
                        [start['lon'], start['lat']],
                        [end['lon'], end['lat']]
                    ]
                },
                'properties': {
                    'pipe_index': i,
                    'sub_type': pipe.get('sub_type', ''),
                    'diameter': pipe.get('diameter', ''),
                    'start_id': pipe['start_id'],
                    'end_id': pipe['end_id']
                }
            })
    return {'type': 'FeatureCollection', 'features': features}



