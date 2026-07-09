"""
Sync pipe data from Excel to PostgreSQL.
Run locally: python sync_pipes.py
"""
import os
import sys
import openpyxl

# Load .env file
env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                os.environ.setdefault(key.strip(), value.strip())

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils.coord_transform import cgcs2000_to_gcj02


def get_db():
    import psycopg2
    db_url = os.environ.get('DATABASE_URL', '')
    if not db_url:
        print("ERROR: DATABASE_URL not set")
        sys.exit(1)
    return psycopg2.connect(db_url)


def init_tables(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS pipe_nodes (
            point_id VARCHAR(50) PRIMARY KEY,
            pipe_type VARCHAR(50),
            sub_type VARCHAR(50),
            feature VARCHAR(100),
            ground_elev FLOAT,
            well_bottom_elev FLOAT,
            depth FLOAT,
            x_cgcs FLOAT,
            y_cgcs FLOAT,
            lon FLOAT,
            lat FLOAT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS pipe_segments (
            id SERIAL PRIMARY KEY,
            pipe_type VARCHAR(50),
            sub_type VARCHAR(50),
            start_id VARCHAR(50),
            end_id VARCHAR(50),
            diameter VARCHAR(50)
        )
    """)
    conn.commit()


def load_nodes(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['管点']
    nodes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        point_id = row[0]
        x_val = row[16]
        y_val = row[17]
        pipe_type = row[9] if len(row) > 9 else None
        sub_type = row[10] if len(row) > 10 else None
        feature = row[15] if len(row) > 15 else None
        ground_elev = row[18] if len(row) > 18 else None
        well_bottom_elev = row[19] if len(row) > 19 else None
        depth = row[11] if len(row) > 11 else None

        if point_id and x_val and y_val:
            try:
                x = float(x_val)
                y = float(y_val)
                lon, lat = cgcs2000_to_gcj02(x, y)

                def safe_float(v):
                    if v is None:
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                nodes.append({
                    'point_id': str(point_id),
                    'pipe_type': str(pipe_type) if pipe_type else '',
                    'sub_type': str(sub_type) if sub_type else '',
                    'feature': str(feature) if feature else '',
                    'ground_elev': safe_float(ground_elev),
                    'well_bottom_elev': safe_float(well_bottom_elev),
                    'depth': safe_float(depth),
                    'x_cgcs': x,
                    'y_cgcs': y,
                    'lon': lon,
                    'lat': lat
                })
            except (ValueError, TypeError):
                continue
    wb.close()
    return nodes


def load_pipes(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['管线']
    pipes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        pipe_type = row[1] if len(row) > 1 else None
        sub_type = row[2] if len(row) > 2 else None
        start_id = row[9] if len(row) > 9 else None
        end_id = row[10] if len(row) > 10 else None
        diameter = row[20] if len(row) > 20 else None

        if start_id and end_id:
            pipes.append({
                'pipe_type': str(pipe_type) if pipe_type else '',
                'sub_type': str(sub_type) if sub_type else '',
                'start_id': str(start_id),
                'end_id': str(end_id),
                'diameter': str(diameter) if diameter else ''
            })
    wb.close()
    return pipes


def sync_to_db(nodes, pipes):
    conn = get_db()
    cur = conn.cursor()

    init_tables(conn)

    # Clear existing data
    cur.execute('DELETE FROM pipe_segments')
    cur.execute('DELETE FROM pipe_nodes')

    # Insert nodes in batches
    batch_size = 100
    for i in range(0, len(nodes), batch_size):
        batch = nodes[i:i+batch_size]
        args = []
        for n in batch:
            args.extend([n['point_id'], n['pipe_type'], n['sub_type'], n['feature'],
                        n['ground_elev'], n['well_bottom_elev'], n['depth'],
                        n['x_cgcs'], n['y_cgcs'], n['lon'], n['lat']])
        placeholders = ','.join(['(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'] * len(batch))
        cur.execute(f"""
            INSERT INTO pipe_nodes (point_id, pipe_type, sub_type, feature, ground_elev, well_bottom_elev, depth, x_cgcs, y_cgcs, lon, lat)
            VALUES {placeholders}
        """, args)
        print(f"  Inserted nodes batch {i//batch_size + 1}")

    # Insert pipes in batches
    for i in range(0, len(pipes), batch_size):
        batch = pipes[i:i+batch_size]
        args = []
        for p in batch:
            args.extend([p['pipe_type'], p['sub_type'], p['start_id'], p['end_id'], p['diameter']])
        placeholders = ','.join(['(%s,%s,%s,%s,%s)'] * len(batch))
        cur.execute(f"""
            INSERT INTO pipe_segments (pipe_type, sub_type, start_id, end_id, diameter)
            VALUES {placeholders}
        """, args)
        print(f"  Inserted pipes batch {i//batch_size + 1}")

    conn.commit()
    conn.close()


if __name__ == '__main__':
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '城西闪传_修复版_去重.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"ERROR: Excel file not found: {xlsx_path}")
        sys.exit(1)

    print("Loading pipe nodes from Excel...")
    nodes = load_nodes(xlsx_path)
    print(f"  Found {len(nodes)} nodes")

    print("Loading pipe segments from Excel...")
    pipes = load_pipes(xlsx_path)
    print(f"  Found {len(pipes)} segments")

    print("Syncing to PostgreSQL...")
    sync_to_db(nodes, pipes)
    print("Done!")
