# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'G:\设备数据分析\城西闪传_修复版_更新后.xlsx', read_only=True)
print('=== Sheets ===')
for sheet_name in wb.sheetnames:
    print(f'  {sheet_name}')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} ===')
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
    print(f'Columns: {len(headers)}')
    for i, h in enumerate(headers):
        print(f'  [{i}] {h}')
    
    row_count = ws.max_row - 1
    print(f'Data rows: {row_count}')
    
    print('Sample data (rows 2-4):')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
        print(f'  Row {i+2}: {list(row)}')
wb.close()
