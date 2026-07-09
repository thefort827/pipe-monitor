import openpyxl
from collections import Counter, defaultdict
import math
import shutil
import os

XLSX_PATH = r'G:\设备数据分析\城西闪传_修复版_更新后.xlsx'
OUTPUT_PATH = r'G:\设备数据分析\城西闪传_清洗后.xlsx'
CLEANED_PATH = r'G:\设备数据分析\城西闪传_清洗后.xlsx'


def load_data():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws_nodes = wb['管点']
    ws_pipes = wb['管线']

    nodes = []
    for i, row in enumerate(ws_nodes.iter_rows(values_only=True)):
        if i < 2:
            nodes.append(row)
            continue
        nodes.append(list(row))

    pipes = []
    for i, row in enumerate(ws_pipes.iter_rows(values_only=True)):
        if i < 2:
            pipes.append(row)
            continue
        pipes.append(list(row))

    return wb, nodes, pipes


def find_duplicate_nodes(nodes):
    point_ids = []
    for i, row in enumerate(nodes):
        if i < 2:
            continue
        point_id = row[0]
        if point_id:
            point_ids.append((str(point_id), i))

    id_counts = Counter([pid for pid, _ in point_ids])
    duplicates = {pid: count for pid, count in id_counts.items() if count > 1}

    dup_details = defaultdict(list)
    for pid, idx in point_ids:
        if pid in duplicates:
            row = nodes[idx]
            x_val = row[16] if len(row) > 16 else None
            y_val = row[17] if len(row) > 17 else None
            dup_details[pid].append({
                'index': idx,
                'x': float(x_val) if x_val else None,
                'y': float(y_val) if y_val else None,
                'row': row
            })

    return dup_details


def find_connected_pipes(pipes, point_id):
    connected = []
    for i, row in enumerate(pipes):
        if i < 2:
            continue
        start = str(row[9]) if len(row) > 9 and row[9] else None
        end = str(row[10]) if len(row) > 10 and row[10] else None
        if start == point_id or end == point_id:
            connected.append(i)
    return connected


def get_pipe_prefix(pipes, point_id):
    prefixes = []
    for i, row in enumerate(pipes):
        if i < 2:
            continue
        start = str(row[9]) if len(row) > 9 and row[9] else None
        end = str(row[10]) if len(row) > 10 and row[10] else None

        if start == point_id:
            other = end
        elif end == point_id:
            other = start
        else:
            continue

        if other and '-' in other:
            prefix = other.split('-')[0]
            if prefix.startswith('WCB') or prefix.startswith('WCN'):
                prefixes.append(prefix)

    if prefixes:
        return Counter(prefixes).most_common(1)[0][0]
    return None


def fix_ysk_duplicates(nodes, pipes):
    print("\n=== 修复 YSK 重复点号 ===")

    ysk_ids = [f'YSK{i}' for i in range(1, 53) if i != 36]

    for ysk_id in ysk_ids:
        indices = []
        for i, row in enumerate(nodes):
            if i < 2:
                continue
            if str(row[0]) == ysk_id:
                indices.append(i)

        if len(indices) <= 1:
            continue

        print(f"\n处理 {ysk_id}（{len(indices)} 个重复）")

        for idx in indices:
            prefix = get_pipe_prefix(pipes, ysk_id)
            if prefix and 'WCB' in prefix:
                suffix = '-A'
            elif prefix and 'WCN' in prefix:
                suffix = '-B'
            else:
                suffix = '-A' if idx == indices[0] else '-B'

            new_id = ysk_id + suffix
            print(f"  行 {idx}: {ysk_id} -> {new_id}")

            nodes[idx][0] = new_id

            for pipe_idx, pipe_row in enumerate(pipes):
                if pipe_idx < 2:
                    continue
                if len(pipe_row) > 9 and str(pipe_row[9]) == ysk_id:
                    pipes[pipe_idx][9] = new_id
                    print(f"    更新管线行 {pipe_idx} 起点号")
                if len(pipe_row) > 10 and str(pipe_row[10]) == ysk_id:
                    pipes[pipe_idx][10] = new_id
                    print(f"    更新管线行 {pipe_idx} 终点号")

    return nodes, pipes


def remove_duplicate_pipes(pipes):
    print("\n=== 删除重复管线 ===")

    seen = {}
    duplicates = []

    for i, row in enumerate(pipes):
        if i < 2:
            continue
        start = str(row[9]) if len(row) > 9 and row[9] else None
        end = str(row[10]) if len(row) > 10 and row[10] else None

        if not start or not end:
            continue

        key = (start, end)
        if key in seen:
            existing_idx = seen[key]
            existing_row = pipes[existing_idx]

            depth1_start = float(existing_row[11]) if len(existing_row) > 11 and existing_row[11] else 0
            depth1_end = float(existing_row[12]) if len(existing_row) > 12 and existing_row[12] else 0
            depth2_start = float(row[11]) if len(row) > 11 and row[11] else 0
            depth2_end = float(row[12]) if len(row) > 12 and row[12] else 0

            score1 = depth1_start + depth1_end
            score2 = depth2_start + depth2_end

            if score2 > score1:
                duplicates.append(existing_idx)
                seen[key] = i
                print(f"  删除行 {existing_idx}（保留行 {i}）: {start} -> {end}")
            else:
                duplicates.append(i)
                print(f"  删除行 {i}（保留行 {existing_idx}）: {start} -> {end}")
        else:
            seen[key] = i

    duplicates.sort(reverse=True)
    for idx in duplicates:
        pipes.pop(idx)

    print(f"  共删除 {len(duplicates)} 条重复管线")
    return pipes


def fix_xfs_w10(nodes):
    print("\n=== 修复 XFS-W10 重复 ===")

    indices = []
    for i, row in enumerate(nodes):
        if i < 2:
            continue
        if str(row[0]) == 'XFS-W10':
            indices.append(i)

    if len(indices) <= 1:
        print("  XFS-W10 无重复")
        return nodes

    print(f"  XFS-W10 有 {len(indices)} 个重复，保留第一个，删除其余")
    for idx in reversed(indices[1:]):
        nodes.pop(idx)

    return nodes


def save_data(wb, nodes, pipes, output_path=None):
    if output_path is None:
        output_path = OUTPUT_PATH

    wb_out = openpyxl.Workbook()

    ws_nodes = wb_out.active
    ws_nodes.title = '管点'
    for row in nodes:
        ws_nodes.append(row)

    ws_pipes = wb_out.create_sheet('管线')
    for row in pipes:
        ws_pipes.append(row)

    wb_out.save(output_path)
    print(f"\n文件已保存到: {output_path}")


def calc_distance(x1, y1, x2, y2):
    return math.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2)


def get_node_coords(nodes, point_id):
    for i, row in enumerate(nodes):
        if i < 2:
            continue
        if str(row[0]) == point_id:
            x = float(row[16]) if len(row) > 16 and row[16] else None
            y = float(row[17]) if len(row) > 17 and row[17] else None
            return x, y
    return None, None


def fix_ysk_pipe_connections(nodes, pipes):
    print("\n=== 修复 YSK 管线连接 ===")

    ysk_b_nodes = {}
    for i, row in enumerate(nodes):
        if i < 2:
            continue
        pid = str(row[0]) if row[0] else ''
        if pid.startswith('YSK') and pid.endswith('-B'):
            try:
                x = float(row[16]) if len(row) > 16 and row[16] else None
                y = float(row[17]) if len(row) > 17 and row[17] else None
            except (ValueError, TypeError):
                continue
            if x and y:
                ysk_b_nodes[pid] = (x, y)

    print(f"  找到 {len(ysk_b_nodes)} 个 YSK-B 节点")

    pipes_to_fix = []
    for pipe_idx, pipe_row in enumerate(pipes):
        if pipe_idx < 2:
            continue

        start = str(pipe_row[9]) if len(pipe_row) > 9 and pipe_row[9] else None
        end = str(pipe_row[10]) if len(pipe_row) > 10 and pipe_row[10] else None

        if not start or not end:
            continue

        ysk_a = None
        wcn_point = None
        pos = None

        if start.startswith('YSK') and start.endswith('-A') and end.startswith('WCN'):
            ysk_a = start
            wcn_point = end
            pos = 'start'
        elif end.startswith('YSK') and end.endswith('-A') and start.startswith('WCN'):
            ysk_a = end
            wcn_point = start
            pos = 'end'

        if ysk_a and wcn_point:
            base_ysk = ysk_a.replace('-A', '')
            ysk_b = base_ysk + '-B'

            if ysk_b in ysk_b_nodes:
                pipes_to_fix.append({
                    'pipe_idx': pipe_idx,
                    'ysk_a': ysk_a,
                    'ysk_b': ysk_b,
                    'wcn_point': wcn_point,
                    'pos': pos
                })

    print(f"  找到 {len(pipes_to_fix)} 条需要修复的管线")

    fixed_count = 0
    for fix_info in pipes_to_fix:
        pipe_idx = fix_info['pipe_idx']
        ysk_a = fix_info['ysk_a']
        ysk_b = fix_info['ysk_b']
        wcn_point = fix_info['wcn_point']
        pos = fix_info['pos']

        if pos == 'start':
            pipes[pipe_idx][9] = ysk_b
        else:
            pipes[pipe_idx][10] = ysk_b

        fixed_count += 1
        print(f"  修复管线行 {pipe_idx}: {ysk_a} -> {ysk_b} (WCN端点: {wcn_point})")

    print(f"  共修复 {fixed_count} 条管线连接")
    return pipes


def main():
    print("=== 数据清洗开始 ===")
    print(f"源文件: {XLSX_PATH}")

    wb, nodes, pipes = load_data()
    print(f"原始管点数: {len(nodes) - 2}")
    print(f"原始管线数: {len(pipes) - 2}")

    dup_nodes = find_duplicate_nodes(nodes)
    print(f"\n发现重复点号: {len(dup_nodes)} 个")
    for pid, details in dup_nodes.items():
        print(f"  {pid}: {len(details)} 个")

    nodes, pipes = fix_ysk_duplicates(nodes, pipes)
    nodes = fix_xfs_w10(nodes)
    pipes = remove_duplicate_pipes(pipes)

    print(f"\n清洗后管点数: {len(nodes) - 2}")
    print(f"清洗后管线数: {len(pipes) - 2}")

    save_data(wb, nodes, pipes)

    print("\n=== 修复 YSK 管线连接 ===")
    print(f"读取文件: {CLEANED_PATH}")

    wb2, nodes2, pipes2 = load_cleaned_data()
    pipes2 = fix_ysk_pipe_connections(nodes2, pipes2)

    save_data(wb2, nodes2, pipes2, CLEANED_PATH)

    print("\n=== 数据清洗完成 ===")


def load_cleaned_data():
    wb = openpyxl.load_workbook(CLEANED_PATH)
    ws_nodes = wb['管点']
    ws_pipes = wb['管线']

    nodes = []
    for i, row in enumerate(ws_nodes.iter_rows(values_only=True)):
        if i < 2:
            nodes.append(row)
            continue
        nodes.append(list(row))

    pipes = []
    for i, row in enumerate(ws_pipes.iter_rows(values_only=True)):
        if i < 2:
            pipes.append(row)
            continue
        pipes.append(list(row))

    return wb, nodes, pipes


if __name__ == '__main__':
    main()
